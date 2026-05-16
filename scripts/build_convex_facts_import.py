from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from word_templates_to_json import (
    build_diagnosis_json,
    build_nursing_json,
    build_pm_json,
    detect_template,
    extract_lines,
)


ROOT = Path(__file__).resolve().parents[1]
CONTENT_DIR = ROOT / "content"
DEFAULT_OUTPUT = CONTENT_DIR / "generated" / "convex_facts.json"

META_KEYS = {
    "module_id",
    "drug_id",
    "dilution_id",
    "type",
    "app",
    "language",
    "title",
    "category",
    "summary",
    "source_note",
    "tags",
}

SECTION_TITLES = {
    "definition": "Definition",
    "orsaker_och_riskfaktorer": "Orsaker och riskfaktorer",
    "symtom": "Symtom",
    "status": "Status",
    "diagnostik": "Diagnostik",
    "behandling": "Behandling",
    "omvardnad": "Omvårdnad",
    "overvakning": "Övervakning",
    "komplikationer": "Komplikationer",
    "eskalering": "Eskalering",
    "praktiska_punkter": "Praktiska punkter",
    "referenser": "Referenser",
    "lakemedel": "Läkemedel",
    "syfte": "Syfte",
    "indikationer": "Indikationer",
    "forberedelser": "Förberedelser",
    "genomforande": "Genomförande",
    "observationer": "Observationer",
    "risker_och_forsiktighet": "Risker och försiktighet",
    "patientinformation": "Patientinformation",
    "dokumentation": "Dokumentation",
    "nar_eskalera": "När eskalera",
}


Fact = dict[str, str]


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def slugify(value: str) -> str:
    value = value.lower().strip()
    value = value.replace("å", "a").replace("ä", "a").replace("ö", "o")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return re.sub(r"-+", "-", value).strip("-")


def normalize(value: str) -> str:
    return slugify(value).replace("-", "")


def summary_from_item(item: dict[str, Any]) -> str:
    summary = item.get("summary")
    if isinstance(summary, dict):
        return clean(summary.get("ingress"))
    return clean(summary)


def section_title(key: str) -> str:
    return SECTION_TITLES.get(key, key.replace("_", " ").capitalize())


def body_from_sections(item: dict[str, Any]) -> str:
    parts: list[str] = []

    for key, value in item.items():
        if key in META_KEYS:
            continue

        if isinstance(value, list) and value:
            lines = [clean(entry) for entry in value if clean(entry)]
            if lines:
                parts.append(section_title(key))
                parts.extend(f"- {line}" for line in lines)
                parts.append("")
        elif isinstance(value, str) and clean(value):
            parts.append(section_title(key))
            parts.append(clean(value))
            parts.append("")

    return "\n".join(parts).strip()


def fact(
    *,
    import_key: str,
    kind: str,
    title: str,
    category: str,
    summary: str,
    body: str,
    source: str,
    status: str = "published",
) -> Fact:
    title = clean(title)
    body = clean(body)
    if not title or not body:
        raise ValueError(f"Kan inte bygga faktapost för {import_key}: titel eller body saknas.")

    return {
        "importKey": import_key,
        "kind": kind,
        "status": status,
        "title": title,
        "category": clean(category) or "Okänd",
        "summary": clean(summary) or body.splitlines()[0][:300],
        "body": body,
        "source": clean(source),
    }


def facts_from_word() -> list[Fact]:
    output: list[Fact] = []

    for path in sorted((CONTENT_DIR / "word").glob("*.docx")):
        lines = extract_lines(path)
        template_type = detect_template(lines, path.name)

        if template_type == "pm":
            item, _ = build_pm_json(lines)
            kind = "pm"
        elif template_type == "nursing":
            item, _ = build_nursing_json(lines)
            kind = "nursing"
        elif template_type == "diagnosis":
            item, _ = build_diagnosis_json(lines)
            kind = "diagnosis"
        else:
            print(f"Hoppar över {path.name}: okänd Word-mall.")
            continue

        output.append(
            fact(
                import_key=f"word:{slugify(path.stem)}",
                kind=kind,
                title=clean(item.get("title")),
                category=clean(item.get("category")),
                summary=summary_from_item(item),
                body=body_from_sections(item),
                source=f"Word: {path.name}",
            )
        )

    return output


def facts_from_json_modules() -> list[Fact]:
    output: list[Fact] = []

    for path in sorted((CONTENT_DIR / "modules").glob("*.json")):
        item = json.loads(path.read_text(encoding="utf-8"))
        body = body_from_sections(item)
        if not body:
            continue

        output.append(
            fact(
                import_key=f"json:module:{clean(item.get('module_id')) or slugify(path.stem)}",
                kind="diagnosis",
                title=clean(item.get("title")),
                category=clean(item.get("category")),
                summary=summary_from_item(item),
                body=body,
                source=clean(item.get("source_note")) or f"JSON: {path.name}",
            )
        )

    return output


def best(current: str, next_value: str) -> str:
    if not current:
        return next_value
    if not next_value:
        return current
    return next_value if len(next_value) > len(current) else current


def row_dict(headers: list[str], row: tuple[Any, ...]) -> dict[str, str]:
    return {header: clean(value) for header, value in zip(headers, row)}


def first_value(row: dict[str, str], aliases: list[str]) -> str:
    normalized = {normalize(key): value for key, value in row.items()}
    for alias in aliases:
        value = normalized.get(normalize(alias), "")
        if value:
            return value
    return ""


def facts_from_drug_excel() -> list[Fact]:
    by_name: dict[str, dict[str, str]] = {}

    for path in sorted((CONTENT_DIR / "drugs").glob("*.xlsx")):
        workbook = load_workbook(path, read_only=True, data_only=True)

        for sheet in workbook.worksheets:
            try:
                headers = [clean(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
            except StopIteration:
                continue
            for raw_row in sheet.iter_rows(min_row=2, values_only=True):
                row = row_dict(headers, raw_row)
                name = first_value(row, ["Läkemedel", "Lakemedel", "Substans/preparat", "Substans"])
                if not name:
                    continue

                key = slugify(name)
                next_record = {
                    "name": name,
                    "group_name": first_value(row, ["Läkemedelsgrupp", "Lakemedelsgrupp", "Grupp"]),
                    "indication": first_value(row, ["Indikation"]),
                    "pharmacodynamics": first_value(row, ["Farmakodynamik"]),
                    "dosage": first_value(row, ["Dosering", "Dos"]),
                    "dilution": first_value(row, ["Spädning", "Spadning"]),
                    "administration": first_value(row, ["Administreringssätt", "Administreringssatt"]),
                    "infusion_time": first_value(row, ["Infusionstid"]),
                    "administration_time": first_value(row, ["Administreringstid"]),
                    "usage_time": first_value(row, ["Användningstid", "Anvandningstid"]),
                    "side_effects": first_value(row, ["Biverkningar"]),
                    "notes": first_value(row, ["Kliniska kommentarer"]),
                    "monitoring_level": first_value(row, ["Övervakningsnivå", "Overvakningsniva"]),
                    "high_risk": first_value(row, ["Högrisk", "Hogrisk"]),
                    "source": first_value(row, ["Källa", "Kalla"]) or path.name,
                }

                current = by_name.get(key, {})
                by_name[key] = {
                    field: best(current.get(field, ""), value)
                    for field, value in next_record.items()
                }

    output: list[Fact] = []
    for key, item in sorted(by_name.items()):
        body_parts = [
            ("Läkemedelsgrupp", item["group_name"]),
            ("Indikation", item["indication"]),
            ("Farmakodynamik", item["pharmacodynamics"]),
            ("Dosering", item["dosage"]),
            ("Spädning", item["dilution"]),
            ("Administrering", item["administration"]),
            ("Infusionstid", item["infusion_time"]),
            ("Administrationstid", item["administration_time"]),
            ("Användningstid", item["usage_time"]),
            ("Biverkningar", item["side_effects"]),
            ("Övervakningsnivå", item["monitoring_level"]),
            ("Högrisk", item["high_risk"]),
            ("Kommentarer", item["notes"]),
        ]
        body = "\n".join(f"{label}: {value}" for label, value in body_parts if value)
        output.append(
            fact(
                import_key=f"excel:drug:{key}",
                kind="medication",
                title=item["name"],
                category=item["group_name"] or "Läkemedel",
                summary=item["indication"] or f"{item['name']} – läkemedelsfakta.",
                body=body,
                source=item["source"],
            )
        )

    return output


def facts_from_dilution_excel() -> list[Fact]:
    output: list[Fact] = []

    for path in sorted((CONTENT_DIR / "dilutions").glob("*.xlsx")):
        workbook = load_workbook(path, read_only=True, data_only=True)

        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            header_index = None
            headers: list[str] = []

            for row_index, raw_row in enumerate(rows):
                candidate = [clean(value) for value in raw_row]
                normalized_headers = [normalize(value) for value in candidate]
                if "substans" in normalized_headers and any(
                    value.startswith("spadningssteg") for value in normalized_headers
                ):
                    header_index = row_index
                    headers = candidate
                    break

            if header_index is None:
                continue

            for index, raw_row in enumerate(rows[header_index + 1:], start=header_index + 2):
                row = row_dict(headers, raw_row)
                name = first_value(row, ["Substans/preparat", "Substans", "Läkemedel", "Lakemedel"])
                if not name or normalize(name) == "substans":
                    continue

                strength = first_value(row, ["Styrka", "Dos/ styrka", "Dos styrka"])
                step_1 = first_value(row, ["Spädningssteg 1", "Spadningssteg 1"])
                step_2 = first_value(row, ["Spädningssteg 2", "Spadningssteg 2"])
                dilution = "\n".join(part for part in [step_1, step_2] if part)
                administration_time = first_value(row, ["Administrerings tid", "Administreringstid"])
                observandum = first_value(row, ["Observandum", "Kommentar"])

                body_parts = [
                    ("Styrka", strength),
                    ("Spädning och dosering", dilution),
                    ("Administreringstid", administration_time),
                    ("Observandum", observandum),
                ]
                body = "\n".join(f"{label}: {value}" for label, value in body_parts if value)
                if not body:
                    continue

                output.append(
                    fact(
                        import_key=f"excel:dilution:{slugify(name)}:{index}",
                        kind="medication",
                        title=f"{name} – spädning",
                        category="Spädning",
                        summary=f"Spädningsfakta för {name}.",
                        body=body,
                        source=f"Excel: {path.name}",
                    )
                )

    return output


def facts_from_cards() -> list[Fact]:
    path = CONTENT_DIR / "cards" / "sepsis_cards.json"
    if not path.exists():
        return []

    cards = json.loads(path.read_text(encoding="utf-8"))
    output: list[Fact] = []
    for card in cards:
        items = [clean(item) for item in card.get("items", []) if clean(item)]
        if not items:
            continue
        output.append(
            fact(
                import_key=f"json:card:{clean(card.get('id')) or slugify(clean(card.get('title')))}",
                kind="card",
                title=clean(card.get("title")),
                category=clean(card.get("category")),
                summary=f"{len(items)} punkter.",
                body="\n".join(f"- {item}" for item in items),
                source=f"JSON: {path.name}",
            )
        )

    return output


def build_facts() -> list[Fact]:
    records = [
        *facts_from_word(),
        *facts_from_json_modules(),
        *facts_from_drug_excel(),
        *facts_from_dilution_excel(),
        *facts_from_cards(),
    ]

    by_key: dict[str, Fact] = {}
    for record in records:
        by_key[record["importKey"]] = record

    return [by_key[key] for key in sorted(by_key)]


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Bygg Convex-faktaimport från SSKHBG:s Word, Excel och JSON-källor."
    )
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    args = parser.parse_args()

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    records = build_facts()
    output_path.write_text(
        json.dumps(records, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    counts: dict[str, int] = {}
    for record in records:
        counts[record["kind"]] = counts.get(record["kind"], 0) + 1

    print(f"KLAR: {output_path}")
    print(f"Faktaposter: {len(records)}")
    for kind, count in sorted(counts.items()):
        print(f"- {kind}: {count}")


if __name__ == "__main__":
    main()
