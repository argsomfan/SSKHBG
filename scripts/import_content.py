import json
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
CONTENT_DIR = BASE_DIR / "content"
MODULES_DIR = CONTENT_DIR / "modules"
DRUGS_DIR = CONTENT_DIR / "drugs"
DILUTIONS_DIR = CONTENT_DIR / "dilutions"
OUTPUT_DIR = BASE_DIR / "src" / "data" / "generated"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def slugify(text):
    if not text:
        return ""

    return (
        str(text)
        .strip()
        .lower()
        .replace("å", "a")
        .replace("ä", "a")
        .replace("ö", "o")
        .replace("/", "-")
        .replace("\\", "-")
        .replace(" ", "-")
        .replace(",", "")
        .replace(".", "")
        .replace("(", "")
        .replace(")", "")
        .replace(":", "")
        .replace("'", "")
        .replace('"', "")
        .replace("µ", "u")
        .replace("+", "plus")
    )


def add_module(
    modules,
    slug,
    title,
    module_type,
    category,
    description,
    source,
    tags,
    sort_order,
):
    modules.append(
        {
            "slug": slug,
            "title": title,
            "type": module_type,
            "category": category,
            "description": description,
            "source": source,
            "last_updated": "2026-03",
            "tags": ", ".join(tags) if isinstance(tags, list) else str(tags or ""),
            "sort_order": sort_order,
        }
    )


def add_section(sections, module_slug, key, title, sort_order):
    sections.append(
        {
            "module_slug": module_slug,
            "key": key,
            "title": title,
            "sort_order": sort_order,
        }
    )


def add_card(
    cards,
    module_slug,
    section_key,
    block_type,
    title,
    body,
    sort_order,
    extra_json=None,
):
    cards.append(
        {
            "module_slug": module_slug,
            "section_key": section_key,
            "block_type": block_type,
            "title": title,
            "body": body,
            "extra_json": extra_json,
            "sort_order": sort_order,
        }
    )


def add_search_terms(search_index, module_slug, terms, default_weight=5):
    if not terms:
        return

    for term in terms:
        if term and str(term).strip():
            search_index.append(
                {
                    "module_slug": module_slug,
                    "term": str(term).strip(),
                    "weight": default_weight,
                }
            )


def import_json_modules():
    modules = []
    sections = []
    cards = []
    relations = []
    search_index = []

    if not MODULES_DIR.exists():
        return modules, sections, cards, relations, search_index

    json_files = sorted(MODULES_DIR.glob("*.json"))
    sort_order = 1

    for file_path in json_files:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        title = data.get("title", "Untitled module")
        slug = data.get("module_id") or slugify(title)
        subtitle = data.get("subtitle", "")
        summary = data.get("summary", {})
        description = subtitle or summary.get("ingress", "")
        source = data.get("source_note", "Imported JSON module")
        tags = data.get("tags", [])

        specialty = data.get("specialty", [])
        category = ", ".join(specialty) if specialty else "PM"

        add_module(
            modules,
            slug=slug,
            title=title,
            module_type="pm",
            category=category,
            description=description,
            source=source,
            tags=tags,
            sort_order=sort_order,
        )
        sort_order += 1

        section_order = 1

        summary = data.get("summary", {})
        if summary:
            add_section(sections, slug, "summary", "Sammanfattning", section_order)
            section_order += 1

            ingress = summary.get("ingress")
            if ingress:
                add_card(cards, slug, "summary", "text", None, ingress, 1)

            card_order = 2
            for goal in summary.get("goals", []):
                add_card(cards, slug, "summary", "bullet", None, goal, card_order)
                card_order += 1

        quick_facts = data.get("quick_facts", [])
        if quick_facts:
            add_section(sections, slug, "quick_facts", "Snabbfakta", section_order)
            section_order += 1

            for idx, item in enumerate(quick_facts, start=1):
                add_card(cards, slug, "quick_facts", "bullet", item.get("label"), item.get("value"), idx)

        immediate_actions = data.get("immediate_actions", [])
        if immediate_actions:
            add_section(sections, slug, "initial_handlaggning", "Initial handläggning", section_order)
            section_order += 1

            for idx, item in enumerate(immediate_actions, start=1):
                add_card(cards, slug, "initial_handlaggning", "step", "Steg %s" % idx, item, idx)

        add_search_terms(search_index, slug, [title], default_weight=10)
        add_search_terms(search_index, slug, data.get("search_terms", []), default_weight=8)
        add_search_terms(search_index, slug, data.get("tags", []), default_weight=6)

    return modules, sections, cards, relations, search_index


def first_value(row_data, possible_keys):
    for key in possible_keys:
        value = row_data.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def import_drugs_master_excel():
    modules = []
    sections = []
    cards = []
    relations = []
    search_index = []

    if not DRUGS_DIR.exists():
        return modules, sections, cards, relations, search_index

    excel_files = sorted(DRUGS_DIR.glob("*.xlsx"))
    sort_order = 1000

    for file_path in excel_files:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb[wb.sheetnames[0]]

        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip() if h else "" for h in header_row]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))

            name = first_value(row_data, ["Läkemedel", "Lakemedel", "Substans/preparat", "Substans", "preparat"])
            if not name:
                continue

            slug = slugify(name)
            group_name = first_value(row_data, ["Läkemedelsgrupp", "Lakemedelsgrupp", "Grupp"])
            indication = first_value(row_data, ["Indikation"])
            pharmacodynamics = first_value(row_data, ["Farmakodynamik"])
            dosage = first_value(row_data, ["Dosering", "Dos"])
            dilution = first_value(row_data, ["Spädning", "Spadning"])
            admin_route = first_value(row_data, ["Administreringssätt", "Administreringssatt"])
            infusion_time = first_value(row_data, ["Infusionstid"])
            administration_time = first_value(row_data, ["Administreringstid"])
            usage_time = first_value(row_data, ["Användningstid", "Anvandningstid"])
            side_effects = first_value(row_data, ["Biverkningar"])
            clinical_comments = first_value(row_data, ["Kliniska kommentarer"])
            monitoring_level = first_value(row_data, ["Övervakningsnivå", "Overvakningsniva"])
            high_risk = first_value(row_data, ["Högrisk", "Hogrisk"])
            source = first_value(row_data, ["Källa", "Kalla"]) or file_path.name

            description_parts = []
            if indication:
                description_parts.append(indication)
            if group_name:
                description_parts.append(group_name)

            description = " – ".join(description_parts) if description_parts else f"{name} – importerat läkemedelskort."

            add_module(
                modules,
                slug=slug,
                title=name,
                module_type="drug",
                category=group_name or "Läkemedel",
                description=description,
                source=source,
                tags=[name, group_name, indication],
                sort_order=sort_order,
            )
            sort_order += 1

            add_section(sections, slug, "indikation", "Indikation", 1)
            add_section(sections, slug, "farmakologi", "Farmakologi", 2)
            add_section(sections, slug, "dosering", "Dosering / Spädning", 3)
            add_section(sections, slug, "administrering", "Administrering", 4)
            add_section(sections, slug, "observandum", "Observandum", 5)

            add_card(cards, slug, "indikation", "text", "Indikation", indication or None, 1)
            add_card(cards, slug, "farmakologi", "text", "Farmakodynamik", pharmacodynamics or None, 1)

            dosage_body_parts = []
            if dosage:
                dosage_body_parts.append("Dosering: " + dosage)
            if dilution:
                dosage_body_parts.append("Spädning: " + dilution)
            add_card(cards, slug, "dosering", "drug_card", None, "\n".join(dosage_body_parts) if dosage_body_parts else None, 1)

            admin_body_parts = []
            if admin_route:
                admin_body_parts.append("Administreringssätt: " + admin_route)
            if infusion_time:
                admin_body_parts.append("Infusionstid: " + infusion_time)
            if administration_time:
                admin_body_parts.append("Administreringstid: " + administration_time)
            if usage_time:
                admin_body_parts.append("Användningstid: " + usage_time)
            add_card(cards, slug, "administrering", "text", None, "\n".join(admin_body_parts) if admin_body_parts else None, 1)

            obs_body_parts = []
            if side_effects:
                obs_body_parts.append("Biverkningar: " + side_effects)
            if clinical_comments:
                obs_body_parts.append("Kliniska kommentarer: " + clinical_comments)
            if monitoring_level:
                obs_body_parts.append("Övervakningsnivå: " + monitoring_level)
            if high_risk:
                obs_body_parts.append("Högrisk: " + high_risk)
            add_card(cards, slug, "observandum", "nursing", None, "\n".join(obs_body_parts) if obs_body_parts else None, 1)

            add_search_terms(
                search_index,
                slug,
                [name, group_name, indication, monitoring_level],
                default_weight=8,
            )

    return modules, sections, cards, relations, search_index


def import_dilutions_excel():
    modules = []
    sections = []
    cards = []
    relations = []
    search_index = []

    if not DILUTIONS_DIR.exists():
        return modules, sections, cards, relations, search_index

    excel_files = sorted(DILUTIONS_DIR.glob("*.xlsx"))
    sort_order = 2000

    for file_path in excel_files:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb[wb.sheetnames[0]]

        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip() if h else "" for h in header_row]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))

            name = first_value(row_data, ["Substans/preparat", "Substans", "Läkemedel", "Lakemedel"])
            if not name:
                continue

            slug = slugify(name + "-spadning")
            strength = first_value(row_data, ["Styrka"])
            dilution = first_value(row_data, ["Spädning och dosering", "Spädning", "Spadning"])
            administration_time = first_value(row_data, ["Administreringstid"])
            observandum = first_value(row_data, ["Observandum"])
            source = file_path.name

            add_module(
                modules,
                slug=slug,
                title=name + " – spädning",
                module_type="drug",
                category="Spädning",
                description=f"Spädningskort för {name}.",
                source=source,
                tags=[name, strength, "spädning"],
                sort_order=sort_order,
            )
            sort_order += 1

            add_section(sections, slug, "indikation", "Översikt", 1)
            add_section(sections, slug, "dosering", "Spädning / Dosering", 2)
            add_section(sections, slug, "administrering", "Administrering", 3)
            add_section(sections, slug, "observandum", "Observandum", 4)

            add_card(cards, slug, "indikation", "text", None, f"Styrka: {strength}" if strength else None, 1)
            add_card(cards, slug, "dosering", "drug_card", "Spädning och dosering", dilution or None, 1)
            add_card(cards, slug, "administrering", "text", "Administreringstid", administration_time or None, 1)
            add_card(cards, slug, "observandum", "nursing", "Observandum", observandum or None, 1)

            add_search_terms(search_index, slug, [name, strength, "spädning"], default_weight=8)

    return modules, sections, cards, relations, search_index


def write_json(filename, data):
    path = OUTPUT_DIR / filename
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def dedupe_dicts(items):
    seen = set()
    output = []

    for item in items:
        key = json.dumps(item, sort_keys=True, ensure_ascii=False)
        if key not in seen:
            seen.add(key)
            output.append(item)

    return output


def main():
    all_modules = []
    all_sections = []
    all_cards = []
    all_relations = []
    all_search_index = []

    importers = [
        import_json_modules,
        import_drugs_master_excel,
        import_dilutions_excel,
    ]

    for importer in importers:
        modules, sections, cards, relations, search_index = importer()
        all_modules.extend(modules)
        all_sections.extend(sections)
        all_cards.extend(cards)
        all_relations.extend(relations)
        all_search_index.extend(search_index)

    all_modules = dedupe_dicts(all_modules)
    all_sections = dedupe_dicts(all_sections)
    all_cards = dedupe_dicts(all_cards)
    all_relations = dedupe_dicts(all_relations)
    all_search_index = dedupe_dicts(all_search_index)

    write_json("modules.json", all_modules)
    write_json("sections.json", all_sections)
    write_json("cards.json", all_cards)
    write_json("relations.json", all_relations)
    write_json("search_index.json", all_search_index)

    print("Import klart.")
    print("Modules:", len(all_modules))
    print("Sections:", len(all_sections))
    print("Cards:", len(all_cards))
    print("Relations:", len(all_relations))
    print("Search index:", len(all_search_index))


if __name__ == "__main__":
    main()